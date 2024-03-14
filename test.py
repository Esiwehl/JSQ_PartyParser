from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook and select the specific sheet
workbook = load_workbook('Bandjes/Wo_Bandjes.xlsx')
sheet = workbook.active

# Dictionary to hold your color labels and their corresponding hex values
# color_labels_to_hex = {}

# for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
#     for cell in row:
#         if cell.value is not None and cell.fill.start_color.type == 'rgb':
#             # Get the RGB value of the fill color
#             argb_value = cell.fill.start_color.rgb
#             # Ensure that argb_value is a string
#             hex_color = f'#{argb_value[2:]}'
#             if isinstance(argb_value, str):
#                 hex_color = f'#{argb_value[2:]}'
#             else:
#                 hex_color = f'#{str(argb_value)[2:]}'
#             color_labels_to_hex[cell.value] = hex_color
cell_color_hex = {}

for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=2, max_col=2):
    for cell in row:
        if cell.fill.start_color.type == 'rgb':
            # Get the ARGB value of the fill color
            argb_value = cell.fill.start_color.rgb
            # Ensure that argb_value is a string
            if isinstance(argb_value, str):
                hex_color = f'#{argb_value[2:]}'
            else:
                f'#{str(argb_value)[2:]}'
            # Store the hex color value with the cell position as the key
            cell_color_hex[f'{cell.column_letter}{cell.row}'] = hex_color


for cell_pos, hex_val in cell_color_hex.items():
    print(f'Cell {cell_pos} has color {hex_val}')
# # Print the color label to hex value mappings
# print(color_labels_to_hex)
