import datetime
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta

parser = argparse.ArgumentParser(description='Transform booking CSV file for JSQ')
parser.add_argument('csv_file', type=str, help='The path to the CSV file to transform')
file = parser.parse_args()
df = pd.read_csv(file.csv_file)

# Define a function to extract the type of party and its duration from the 'Items' column
def getPartyType(items):
    if 'Snack' in items:
        party_type = 'Snack'
    elif 'Meal' in items:
        party_type = 'Meal'
    else:
        party_type = 'Onbekend'
    
    # Extracting the duration
    if '60 min' in items:
        duration = 60
    elif '90 min' in items:
        duration = 90
    elif '120 min' in items:
        duration = 120
    else:
        duration = None

    return f'{party_type} {duration}' if duration else party_type

# Define a function to extract notes from the 'Items' column
def getNotes(items):
    notes_keywords = ['slush', 'taart', 'diploma', 'snoepzakjes', 'bracelet']
    notes = [note for note in notes_keywords if note in items.lower()]
    return ', '.join(notes)

# Define a function to calculate the arrival time as 30 minutes before session start
def calculate_time(time_str):
    time = datetime.strptime(time_str, '%I:%M %p')
    new_time = time - timedelta(minutes=30)
    return new_time.strftime('%H:%M')

# Create a new DataFrame
transformed_data = {
    'Binnenkomst': [],
    'Achternaam': [],
    'Opmerkingen': [],
    'Arrangement': [],
    'Aantal': [],
    'Eettijd': [],
    'Tafelnummer':[]
}

for _, row in df.iterrows():
    transformed_data['Binnenkomst'].append(calculate_time(row['Session Start Time']))
    transformed_data['Achternaam'].append(row['Last Name'])
    transformed_data['Opmerkingen'].append(getNotes(row['Items']))
    transformed_data['Arrangement'].append(getPartyType(row['Items']))
    transformed_data['Aantal'].append(row['Guests'])
    # Set the eating time for meal parties, equal to 'Session End Time'
    # calculate_time now necessary here because Session end Time now relates to booking end instead of jump session end.
    transformed_data['Eettijd'].append(calculate_time(row['Session End Time'])if 'Meal' in row['Items'] else '-')
    transformed_data['Tafelnummer'].append('')

# Convert the transformed data into a DataFrame
transformed_df = pd.DataFrame(transformed_data)
transformed_df.sort_values(by='Binnenkomst', inplace=True)

expanded_data = {col: [] for col in transformed_df.columns}
for _, row in transformed_df.iterrows():
    for col in transformed_df.columns:
        expanded_data[col].append(row[col])
    # Add empty row
    for col in transformed_df.columns:
        expanded_data[col].append('')

expanded_df = pd.DataFrame(expanded_data)
# booking_date = pd.to_datetime(df['Booking Date'].iloc[0].strftime)
df['Day of Week'] = pd.to_datetime(df['Booking Date']).dt.day_name()
df['Day of Month'] = pd.to_datetime(df['Booking Date']).dt.day

booking_date_str = pd.to_datetime(df['Booking Date'].iloc[0]).strftime('%Y-%m-%d')
day_of_week = pd.to_datetime(df['Booking Date'].iloc[0]).strftime('%A')[:3]
file_day_of_month = df['Day of Month'].iloc[0]

# Create an Excel writer object and write the DataFrame
reformatted_file = f"{day_of_week}_{file_day_of_month}_FormattedForm.xlsx"

writer = pd.ExcelWriter(reformatted_file, engine='openpyxl')


# Write the DataFrame without the index
expanded_df.to_excel(writer, index=False, sheet_name='Bookings')

# Save the Excel file to access the workbook and worksheet objects
writer.close()

# Load the workbook and worksheet for formatting
workbook = openpyxl.load_workbook(reformatted_file)
worksheet = workbook['Bookings']

# Define the font and size
font = Font(name='Times New Roman', size=18)

# Apply the font to all cells that are not empty
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value is not None and cell.value != '':
            cell.font = font

uniform_column_width = 20.71
for column in worksheet.columns:
    worksheet.column_dimensions[column[0].column_letter].width = uniform_column_width

# Set a uniform row height (in points)
uniform_row_height = 22.5
for row in worksheet.iter_rows():
    for cell in row:
        worksheet.row_dimensions[cell.row].height = uniform_row_height


# Save the workbook with the formatting
workbook.save(reformatted_file)
workbook.close()

print(f"Nieuw Excel bestand: {reformatted_file}")

# Show the transformed DataFrame to verify correctness
transformed_df.head()
